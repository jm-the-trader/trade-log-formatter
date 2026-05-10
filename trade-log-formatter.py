import pandas as pd
import re
import json
import os
from datetime import datetime, timedelta
from glob import glob
import pymupdf as fitz  # PyMuPDF >= 1.25 uses pymupdf import
import csv

# Configuration
DEBUG = False  # Set to False to enable debug printing
# Set default test date to yesterday
yesterday = datetime.now() - timedelta(days=1)
# DEFAULT_TEST_DATE = datetime.now().strftime("%m.%Y")

# Add at the top with other configurations
TEST_MODE = False  # Set to True to use test files
DEFAULT_TEST_DATE = "06.2025" if TEST_MODE else datetime.now().strftime("%m.%Y")
MASTER_FILE = "master-copy-test.xlsx" if TEST_MODE else "master-trades.xlsx"
MASTER_BACKUP = "master-copy-test-backup.xlsx" if TEST_MODE else "master-copy-backup.xlsx"
PROCESSED_FILE = "processed_files_test.json" if TEST_MODE else "processed_files.json"
BASE_PATH_TRADES = "/Users/michaeljacinto/Library/CloudStorage/OneDrive-Personal/Desktop - onedrive/trades"

def debug_print(*args, **kwargs):
    """Wrapper for debug printing"""
    if DEBUG:
        print(*args, **kwargs)

def get_folder_path(date_str):
    """Find folder containing the input month-year"""
    try:
        # Parse input date string (e.g., 05.2025)
        target_date = datetime.strptime(date_str, "%m.%Y")
        target_folder = target_date.strftime("%m.%Y")
        
        # Look for exact month folder
        folder_path = os.path.join(BASE_PATH_TRADES, target_folder)
        
        if os.path.exists(folder_path) and os.path.isdir(folder_path):
            debug_print(f"Found matching folder: {target_folder}")
            return folder_path
            
        raise FileNotFoundError(f"No folder found for month-year: {date_str}")
            
    except ValueError as e:
        print(f"⚠️ Invalid date format: {e}. Expected MM.YYYY (e.g., 05.2025)")
        raise ValueError("Invalid date format. Please use MM.YYYY (e.g., 05.2025)")


def parse_trade_line(line):
    """Parse a single trade line from PDF report"""
    # Updated pattern to handle options symbols
    pattern = re.compile(r"""
        U\*\*\*\d+\s+               # Account ID (masked)
        (?P<symbol>[A-Z\s\d]+)\s+   # Symbol (including options)
        (?P<trade_date>\d{4}-\d{2}-\d{2}),?\s*  # Trade Date (optional comma)
        (?P<trade_time>\d{2}:\d{2}:\d{2})\s*    # Trade Time
        (?P<settle_date>\d{4}-\d{2}-\d{2})\s*   # Settle Date
        [-\s]*                      # Exchange separator
        (?P<type>BUY|SELL)\s*      # Trade Type
        (?P<quantity>-?\d+)\s*      # Quantity (allowing negative numbers)
        (?P<price>\d+\.?\d*)\s*     # Price
        [-\d.,\s]*                  # Proceeds
    """, re.VERBOSE | re.IGNORECASE)

    match = pattern.search(line)
    if not match:
        # Analyze why the pattern failed to match
        checks = [
            ("Account ID", r"U\*\*\*\d+"),
            ("Symbol", r"[A-Z\s\d]+"),
            ("Trade Date", r"\d{4}-\d{2}-\d{2}"),
            ("Time", r"\d{2}:\d{2}:\d{2}"),
            ("Trade Type", r"BUY|SELL"),
            ("Quantity", r"-?\d+"),
            ("Price", r"\d+\.?\d*")
        ]
        
        print("\n  🔍 Pattern match failure analysis:")
        for check_name, check_pattern in checks:
            if not re.search(check_pattern, line):
                print(f"    ❌ Missing {check_name}")
        print(f"    📝 Raw text: {line[:100]}...")
        return None

    trade_data = {
        "Symbol": match.group("symbol").strip(),  # Strip extra whitespace
        "Date": match.group("trade_date"),
        "Time": match.group("trade_time"),
        "Quantity": abs(int(match.group("quantity"))),  # Use absolute value
        "Price": float(match.group("price")),
        "Side": match.group("type").upper()
    }
    
    return trade_data

def is_option_trade(symbol):
    """Check if the trade is an options trade by looking for date pattern after symbol"""
    # Match pattern like: UNH 16JAN26 550 C
    return bool(re.search(r'[A-Z]+\s+\d+[A-Z]{3}\d{2}\s+\d+\s+[CP]', symbol))

def extract_trades_from_pdf(file_path):
    """Extract all trades from a PDF file and show summary"""
    trades = []
    try:
        doc = fitz.open(file_path)
        print(f"\n📄 Processing: {os.path.basename(file_path)} ({len(doc)} pages)")
        
        # Process ALL pages
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            debug_print(f"  📄 Processing page {page_num + 1}")
            
            # Check if this page has any trade data
            if 'U***' not in text:
                debug_print(f"    ⏭️ No trade data found on page {page_num + 1}")
                continue
            
            # For first page, look for USD sections
            # For subsequent pages, process the entire page as continuation
            if page_num == 0:
                # Look for USD sections on first page only
                sections = text.split('USD')
                
                if len(sections) <= 1:
                    debug_print(f"    ⏭️ No USD sections found on page {page_num + 1}")
                    continue
                
                # Process each USD section (skip the first split which is before first USD)
                sections_to_process = sections[1:]
            else:
                # For continuation pages, process the entire page content
                debug_print(f"    📋 Processing continuation page {page_num + 1}")
                sections_to_process = [text]  # Process entire page as one section
            
            # Process sections
            for section_num, section in enumerate(sections_to_process, 1):
                if page_num == 0:
                    debug_print(f"    📋 Processing USD section {section_num} on page {page_num + 1}")
                else:
                    debug_print(f"    📋 Processing continuation content on page {page_num + 1}")
                
                # End processing at Financial Instrument Information if found
                if 'Financial Instrument Information' in section:
                    relevant_text = section.split('Financial Instrument Information')[0]
                else:
                    relevant_text = section
                
                lines = [line.strip() for line in relevant_text.splitlines() if line.strip()]
                
                if not lines:
                    debug_print(f"      ⏭️ No lines found in section")
                    continue
                
                i = 0
                trades_found_in_section = 0
                while i < len(lines):
                    if lines[i].startswith('U***'):
                        try:
                            # Make sure we have enough lines
                            if i + 7 >= len(lines):
                                debug_print(f"      ⚠️ Not enough lines for trade at line {i}")
                                i += 1
                                continue
                                
                            # Extract trade data from lines
                            account = lines[i]
                            symbol = lines[i+1]     # This might be an option symbol
                            datetime = lines[i+2]
                            trade_type = lines[i+5].strip().upper()
                            quantity = lines[i+6]
                            price = lines[i+7]
                            
                            # Skip if this is a Total line
                            if "Total" not in symbol:
                                # Keep full symbol if it's an option
                                is_option = is_option_trade(symbol)
                                trade_symbol = symbol if is_option else symbol.split()[0]
                                
                                # For options, multiply price by 100
                                raw_price = float(price.strip())
                                adjusted_price = raw_price * 100 if is_option else raw_price
                                
                                trade_data = {
                                    "Symbol": trade_symbol,
                                    "Date": datetime.split(',')[0],
                                    "Time": datetime.split(',')[1].strip(),
                                    "Quantity": int(quantity.strip()),
                                    "Price": adjusted_price,
                                    "Side": trade_type
                                }
                                
                                debug_print(f"      ✅ Parsed Trade: {'LONG' if trade_data['Side'] == 'BUY' else 'SHORT'} {trade_data['Quantity']} "
                                          f"{trade_data['Symbol']} @ ${trade_data['Price']:.2f} "
                                          f"({'Option' if is_option else 'Stock'}) [Page {page_num + 1}]")
                                
                                trades.append(trade_data)
                                trades_found_in_section += 1
                            
                            # Skip to next potential transaction
                            i += 12
                        except (IndexError, ValueError) as e:
                            debug_print(f"      ⚠️ Error parsing trade at line {i} on page {page_num + 1}")
                            debug_print(f"      ⚠️ Error details: {str(e)}")
                            debug_print(f"      ⚠️ Current line content: {lines[i] if i < len(lines) else 'EOF'}")
                            i += 1
                    else:
                        i += 1
                
                debug_print(f"    📊 Found {trades_found_in_section} trades on page {page_num + 1}")
        
        # Add summary at the end of each PDF
        if DEBUG:
            print("\n  📊 Debug Summary of Trades:")
            if trades:
                # Group trades by symbol and side
                buys = {}
                sells = {}
                
                # Debug the trade sorting
                debug_print("\n  🔍 Sorting trades:")
                for trade in trades:
                    symbol = trade['Symbol']
                    side = trade['Side']
                    debug_print(f"    Trade: {symbol} {side} {trade['Quantity']} @ {trade['Price']}")
                    
                    # Determine target dictionary based on trade side
                    if side == "SELL":
                        target_dict = sells
                    else:
                        target_dict = buys
                    
                    if symbol not in target_dict:
                        target_dict[symbol] = {
                            'qty': 0,
                            'total_cost': 0,
                            'earliest_time': trade['Time']
                        }
                    
                    current = target_dict[symbol]
                    current['qty'] += trade['Quantity']
                    current['total_cost'] += trade['Quantity'] * trade['Price']
                    current['earliest_time'] = min(current['earliest_time'], trade['Time'])
                    
                    debug_print(f"    Added to {'SELLS' if side == 'SELL' else 'BUYS'}, "
                            f"New total: {current['qty']} @ {current['total_cost']/current['qty']:.2f}")
                
                # Print summary
                print("\n  📊 PDF Summary:")
                pdf_total = 0
                
                # Print LONG summary
                if buys:
                    print("\n  🟢 LONG:")  # Changed from BUYS
                    print("  Symbol  Shares    Avg Price    Total Value    Time")
                    print("  " + "-" * 55)
                    
                    for symbol, data in buys.items():
                        if data['qty'] > 0:
                            avg_price = data['total_cost'] / data['qty']
                            total_value = data['total_cost']
                            pdf_total += total_value
                            
                            print(f"  {symbol:6} {data['qty']:8.0f} @ ${avg_price:8,.2f} = ${total_value:11,.2f}  {data['earliest_time']}")
                    
                    print("  " + "-" * 55)
                
                # Print SHORT summary
                if sells:
                    print("\n  🔴 SHORT:")  # Changed from SELLS
                    print("  Symbol  Shares    Avg Price    Total Value    Time")
                    print("  " + "-" * 55)
                    
                    for symbol, data in sells.items():
                        if data['qty'] < 0:
                            avg_price = data['total_cost'] / data['qty']
                            total_value = data['total_cost']
                            pdf_total += total_value
                            
                            print(f"  {symbol:6} {data['qty']:8.0f} @ ${avg_price:8,.2f} = ${total_value:11,.2f}  {data['earliest_time']}")
                    
                    print("  " + "-" * 55)
                
                print(f"  PDF Total Value: ${pdf_total:,.2f}\n")
        
        doc.close()
        
    except Exception as e:
        print(f"❌ Error processing {file_path}: {str(e)}")
    
    return trades

def gather_all_trades(folder):
    """Gather trades from all PDFs in chronological order"""
    all_trades = []
    # Get all PDF files and sort them by date in filename
    pdf_files = glob(os.path.join(folder, "DailyTradeReport.*.pdf"))
    
    # Sort PDFs by date in filename (format: DailyTradeReport.YYYYMMDD.pdf)
    pdf_files.sort(key=lambda x: os.path.basename(x).split('.')[1])
    
    processed_files = manage_processed_files(folder, check_only=True)
    
    new_files = False
    for pdf in pdf_files:
        filename = os.path.basename(pdf)
        if filename in processed_files:
            debug_print(f"⏭️  Skipping previously processed file: {filename}")
            continue
            
        new_files = True
        debug_print(f"\n📅 Processing {filename}")
        trades = extract_trades_from_pdf(pdf)
        all_trades.extend(trades)
        
        # Mark file as processed
        manage_processed_files(folder, filename)
    
    if not new_files:
        print("\n📝 No new trade reports to process")
    
    return all_trades

def export_to_csv(trades, output_file, folder_path):
    """Export trades to CSV file in the same folder as PDFs"""
    if not trades:
        print("No trades found to export.")
        return

    fields = ["Symbol", "Quantity", "Side", "Price", "Time", "Date"]
    
    # Create full path for output file in the same folder as PDFs
    output_path = os.path.join(folder_path, output_file)
    
    with open(output_path, mode='w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(trades)

    print(f"✅ Exported {len(trades)} trades to {output_path}")

def consolidate_trades(trades):
    """Consolidate trades by symbol and date, averaging prices for same-day trades"""
    consolidated = {}
    
    for trade in trades:
        # Update trade side before creating key
        trade['Side'] = 'LONG' if trade['Side'] == 'BUY' else 'SHORT'
        key = (trade['Symbol'], trade['Date'], trade['Side'])
        
        if key in consolidated:
            existing = consolidated[key]
            # Calculate new total quantity and weighted average price
            total_qty = existing['Quantity'] + trade['Quantity']
            weighted_price = (
                (existing['Quantity'] * existing['Price'] + 
                 trade['Quantity'] * trade['Price']) / total_qty
            )
            
            # For SHORT orders, keep the latest time
            # For LONG orders, keep the earliest time
            if trade['Side'] == 'SHORT':
                time_to_use = max(existing['Time'], trade['Time'])
            else:
                time_to_use = min(existing['Time'], trade['Time'])
            
            consolidated[key] = {
                'Symbol': trade['Symbol'],
                'Date': trade['Date'],
                'Time': time_to_use,
                'Side': trade['Side'],
                'Quantity': total_qty,
                'Price': weighted_price
            }
        else:
            consolidated[key] = trade.copy()
    
    return list(consolidated.values())

def check_open_positions(folder_path):
    """Check master copy for open positions and provide summary with totals"""
    try:
        master_file = os.path.join(BASE_PATH_TRADES, MASTER_FILE)
        df = pd.read_excel(master_file)
        
        # Find rows where Exit Qty or Exit Price is empty/NaN
        open_positions = df[df['Exit Qty'].isna() | df['Exit Price'].isna()]
        
        if not open_positions.empty:
            print("\n📈 Open Positions (Detail):")
            for _, row in open_positions.iterrows():
                position_type = "LONG" if row['Side'] in ['BUY', 'LONG'] else "SHORT"
                print(f"  • {row['Symbol']}: {row['Qty']} shares ({position_type}) @ ${row['Entry Price']:.2f} "
                      f"({row['Entry Date']} {row['Entry Time']})")
            
            # Create summary by symbol
            summary = {}
            grand_total = 0
            
            for _, row in open_positions.iterrows():
                symbol = row['Symbol']
                qty = row['Qty']
                price = row['Entry Price']
                date = pd.to_datetime(row['Entry Date'])
                
                if symbol in summary:
                    existing = summary[symbol]
                    total_qty = existing['qty'] + qty
                    weighted_price = (existing['qty'] * existing['price'] + qty * price) / total_qty
                    earliest_date = min(existing['date'], date)
                    
                    summary[symbol] = {
                        'qty': total_qty,
                        'price': weighted_price,
                        'date': earliest_date,
                        'total_value': total_qty * weighted_price
                    }
                else:
                    summary[symbol] = {
                        'qty': qty,
                        'price': price,
                        'date': date,
                        'total_value': qty * price
                    }
            
            # Print summary with position values
            print("\n📊 Open Positions Summary:")
            print("  Symbol  Shares    Avg Price    Total Value    Since")
            print("  " + "-" * 55)
            
            for symbol, data in summary.items():
                position_value = data['total_value']
                grand_total += position_value
                print(f"  {symbol:6} {data['qty']:8.0f} @ ${data['price']:8,.2f} = ${position_value:11,.2f}  {data['date'].strftime('%Y-%m-%d')}")
            
            print("  " + "-" * 55)
            print(f"  Total Portfolio Value: ${grand_total:,.2f}")
            
            return open_positions.to_dict('records')
        else:
            print("\n✅ No open positions found")
            return []
            
    except FileNotFoundError:
        print(f"\n⚠️  Master copy not found: {master_file}")
        return []
    except Exception as e:
        print(f"\n❌ Error reading master copy: {str(e)}")
        return []

def _is_blank(v):
    """True if value is None, NaN, or empty string."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == '':
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return False


def _row_state(row):
    """Categorize a master row: 'closed', 'legacy_partial', or 'open'."""
    eq_blank = _is_blank(row.get('Exit Qty'))
    ed_blank = _is_blank(row.get('Exit Date'))
    if not eq_blank and not ed_blank:
        return 'closed'
    if not eq_blank and ed_blank:
        return 'legacy_partial'
    return 'open'


def _match_trades_fifo_records(rows, consolidated_trades):
    """Core FIFO matcher operating on a list of dict records.

    Each input row may carry arbitrary extra keys beyond the FIFO fields;
    those keys are preserved on outputs (used for lineage tracking when
    writing back to user spreadsheets).

    Existing rows fall into three buckets:
      - closed:         Exit Qty + Exit Date both filled. Left untouched.
      - legacy_partial: Exit Qty filled but Exit Date blank. New closes are
                        weighted-averaged into Exit Qty/Price; Exit Date is
                        set to the final fill date once fully closed.
      - open:           Exit Qty blank. Approach B (split-on-close): a partial
                        close splits the row into a closed lot of size `fill`
                        and a remaining open row of size `qty - fill`.

    Any new trade quantity that doesn't match an opposite-side open row
    becomes a new open BUY/SELL row.
    """
    # Normalize 'LONG'/'SHORT' to 'BUY'/'SELL' for matching.
    for r in rows:
        if r.get('Side') == 'LONG':
            r['Side'] = 'BUY'
        elif r.get('Side') == 'SHORT':
            r['Side'] = 'SELL'

    def entry_dt(r):
        try:
            return pd.to_datetime(f"{r.get('Entry Date')} {r.get('Entry Time')}", errors='coerce')
        except Exception:
            return pd.NaT

    from collections import defaultdict
    queues = defaultdict(list)
    indexed = sorted(
        enumerate(rows),
        key=lambda ir: (entry_dt(ir[1]) if not pd.isna(entry_dt(ir[1])) else pd.Timestamp.max),
    )
    for i, r in indexed:
        if _row_state(r) in ('open', 'legacy_partial'):
            queues[r['Symbol']].append(i)

    trades = sorted(consolidated_trades, key=lambda x: (x['Date'], x['Time']))
    print("\n🔄 Matching trades using FIFO method...")

    for trade in trades:
        symbol = trade['Symbol']
        side = trade['Side']  # 'LONG' or 'SHORT' after consolidate_trades
        qty = abs(int(trade['Quantity']))
        price = float(trade['Price'])
        time = trade['Time']
        date = trade['Date']

        if qty == 0:
            continue

        # LONG (BUY) closes open SELL rows; remainder opens BUY.
        # SHORT (SELL) closes open BUY rows; remainder opens SELL.
        match_side = 'SELL' if side == 'LONG' else 'BUY'
        new_row_side = 'BUY' if side == 'LONG' else 'SELL'
        new_row_qty_sign = 1 if new_row_side == 'BUY' else -1
        # Exit Qty has opposite sign of the closed row's Qty.
        exit_qty_sign = -1 if match_side == 'BUY' else 1

        remaining = qty
        q = queues[symbol]
        print(f"\n📊 {side} {qty} {symbol} @ ${price:.2f}")

        i_idx = 0
        while remaining > 0 and i_idx < len(q):
            r = rows[q[i_idx]]
            if r['Side'] != match_side:
                i_idx += 1
                continue

            state = _row_state(r)
            row_qty = abs(int(r['Qty']))

            if state == 'legacy_partial':
                existing_eq = abs(float(r['Exit Qty']))
                existing_ep = float(r['Exit Price'])
                room = row_qty - existing_eq
                if room <= 0:
                    r['Exit Date'] = date
                    if _is_blank(r.get('Exit Time')):
                        r['Exit Time'] = time
                    q.pop(i_idx)
                    continue
                fill = min(remaining, room)
                new_eq = existing_eq + fill
                new_ep = (existing_eq * existing_ep + fill * price) / new_eq
                r['Exit Qty'] = exit_qty_sign * new_eq
                r['Exit Price'] = new_ep
                r['Exit Time'] = time
                if new_eq >= row_qty:
                    r['Exit Date'] = date  # final fill date
                    q.pop(i_idx)
                else:
                    i_idx += 1
                remaining -= fill
                print(f"  → legacy avg-close {fill} ({symbol}) → exit_qty={new_eq}")

            elif state == 'open':
                fill = min(remaining, row_qty)
                if fill >= row_qty:
                    r['Exit Qty'] = exit_qty_sign * fill
                    r['Exit Price'] = price
                    r['Exit Time'] = time
                    r['Exit Date'] = date
                    q.pop(i_idx)
                    print(f"  → closed full lot {fill} {symbol}")
                else:
                    row_sign = 1 if r['Side'] == 'BUY' else -1
                    closed_lot = dict(r)
                    closed_lot['Qty'] = row_sign * fill
                    closed_lot['Exit Qty'] = exit_qty_sign * fill
                    closed_lot['Exit Price'] = price
                    closed_lot['Exit Time'] = time
                    closed_lot['Exit Date'] = date
                    closed_lot['_split_child'] = True  # lineage flag for journal writer
                    rows.append(closed_lot)
                    r['Qty'] = row_sign * (row_qty - fill)
                    print(f"  → split lot {symbol}: closed {fill}, {row_qty - fill} still open")
                remaining -= fill
            else:
                i_idx += 1

        if remaining > 0:
            new_row = {
                'Symbol': symbol,
                'Qty': new_row_qty_sign * remaining,
                'Side': new_row_side,
                'Entry Price': price,
                'Entry Time': time,
                'Entry Date': date,
                'Notes': '',
                'Exit Qty': None,
                'Exit Price': None,
                'Exit Time': None,
                'Exit Date': None,
            }
            new_idx = len(rows)
            rows.append(new_row)
            queues[symbol].append(new_idx)
            print(f"  → opened new {new_row_side} {remaining} {symbol}")

    return rows


def match_trades_fifo(df_master, consolidated_trades):
    """DataFrame wrapper around `_match_trades_fifo_records` for the
    master-trades.xlsx schema. Drops any extra columns to keep the master
    sheet clean.
    """
    columns = [
        "Symbol", "Qty", "Side", "Entry Price", "Entry Time",
        "Entry Date", "Notes", "Exit Qty", "Exit Price",
        "Exit Time", "Exit Date",
    ]

    if df_master is None or df_master.empty:
        rows = []
    else:
        df_clean = df_master[[c for c in columns if c in df_master.columns]].copy()
        for c in columns:
            if c not in df_clean.columns:
                df_clean[c] = None
        rows = df_clean.to_dict('records')

    rows = _match_trades_fifo_records(rows, consolidated_trades)

    df_result = pd.DataFrame(rows, columns=columns)
    if not df_result.empty:
        df_result['_dt'] = pd.to_datetime(
            df_result['Entry Date'].astype(str) + ' ' + df_result['Entry Time'].astype(str),
            format='mixed', errors='coerce',
        )
        df_result = df_result.sort_values('_dt', kind='stable').drop(columns='_dt').reset_index(drop=True)
    return df_result


def update_master_sheet(consolidated_trades, folder_path):
    """Update master balance sheet with new trades after backing up"""
    try:
        BASE_PATH = BASE_PATH_TRADES
        master_file = os.path.join(BASE_PATH, MASTER_FILE)
        backup_file = os.path.join(BASE_PATH, MASTER_BACKUP)
        
        # Create backup of current master file
        if os.path.exists(master_file):
            print(f"\n📑 Creating backup of master sheet...")
            # Read existing workbook with all sheets
            with pd.ExcelFile(master_file) as xls:
                all_sheets = {}
                for sheet_name in xls.sheet_names:
                    all_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
            
            # Save backup with all sheets
            with pd.ExcelWriter(backup_file, engine='openpyxl') as writer:
                for sheet_name, df in all_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Backup created: {os.path.basename(MASTER_BACKUP)}")
            
            # Load existing sheets with proper names - handle both old and new sheet names
            df_master = pd.DataFrame()
            df_raw_trades = pd.DataFrame()
            df_consolidated = pd.DataFrame()
            
            # Get Trades sheet (formerly Sheet1)
            if 'Trades' in all_sheets:
                df_master = all_sheets['Trades'].copy()
            elif 'Sheet1' in all_sheets:
                df_master = all_sheets['Sheet1'].copy()
            else:
                df_master = pd.DataFrame(columns=[
                    "Symbol", "Qty", "Side", "Entry Price", "Entry Time", 
                    "Entry Date", "Notes", "Exit Qty", "Exit Price", 
                    "Exit Time", "Exit Date"
                ])
            
            # Get Raw Trades sheet
            if 'Raw Trades' in all_sheets:
                df_raw_trades = all_sheets['Raw Trades'].copy()
            else:
                df_raw_trades = pd.DataFrame(columns=[
                    "Symbol", "Quantity", "Side", "Price", "Time", "Date"
                ])
        
            # Get Consolidated Trades sheet with correct column order
            if 'Consolidated Trades' in all_sheets:
                df_consolidated = all_sheets['Consolidated Trades'].copy()
            else:
                df_consolidated = pd.DataFrame(columns=[
                    "Symbol", "Quantity", "Side", "Avg_Price", "Time", "Processed"
                ])
                
        else:
            # Create new master file with headers
            df_master = pd.DataFrame(columns=[
                "Symbol", "Qty", "Side", "Entry Price", "Entry Time", 
                "Entry Date", "Notes", "Exit Qty", "Exit Price", 
                "Exit Time", "Exit Date"
            ])
            df_raw_trades = pd.DataFrame(columns=[
                "Symbol", "Quantity", "Side", "Price", "Time", "Date"
            ])
            df_consolidated = pd.DataFrame(columns=[
                "Symbol", "Quantity", "Side", "Avg_Price", "Time", "Processed"
            ])
        
        print(f"📊 Current data before processing:")
        print(f"   - Trades sheet: {len(df_master)} rows")
        print(f"   - Raw Trades sheet: {len(df_raw_trades)} rows")
        print(f"   - Consolidated Trades sheet: {len(df_consolidated)} rows")
        
        # Create unique identifier for existing trades in master sheet
        if not df_master.empty:
            # Check which column names exist and use them
            qty_col = 'Qty' if 'Qty' in df_master.columns else ('Quantity' if 'Quantity' in df_master.columns else None)
            
            # Check for Side column more carefully
            if 'Side' in df_master.columns:
                side_col = 'Side'
            elif 'Type' in df_master.columns:
                side_col = 'Type'
            else:
                print(f"⚠️ Warning: No 'Side' or 'Type' column found in master sheet. Available columns: {list(df_master.columns)}")
                side_col = None
            
            price_col = 'Entry Price' if 'Entry Price' in df_master.columns else ('Price' if 'Price' in df_master.columns else None)
            time_col = 'Entry Time' if 'Entry Time' in df_master.columns else ('Time' if 'Time' in df_master.columns else None)
            date_col = 'Entry Date' if 'Entry Date' in df_master.columns else ('Date' if 'Date' in df_master.columns else None)
            
            # Only create trade keys if we have all required columns
            if all(col is not None for col in [qty_col, side_col, price_col, time_col, date_col]):
                df_master['trade_key'] = (
                    df_master['Symbol'].astype(str) + '_' + 
                    df_master[qty_col].astype(str) + '_' + 
                    df_master[side_col].astype(str) + '_' + 
                    df_master[price_col].astype(str) + '_' + 
                    df_master[time_col].astype(str) + '_' + 
                    df_master[date_col].astype(str)
                )
            else:
                print(f"⚠️ Warning: Missing required columns in master sheet. Skipping trade key creation.")
                df_master['trade_key'] = ''
        
        # Create unique identifier for existing trades in raw trades sheet
        if not df_raw_trades.empty:
            # Check which column names exist and use them - be more careful about fallbacks
            qty_col = 'Quantity' if 'Quantity' in df_raw_trades.columns else 'Qty'
            
            # Check for Side column more carefully
            if 'Side' in df_raw_trades.columns:
                side_col = 'Side'
            elif 'Type' in df_raw_trades.columns:
                side_col = 'Type'
            else:
                # If neither exists, we have a problem with the data structure
                print(f"⚠️ Warning: No 'Side' or 'Type' column found in Raw Trades. Available columns: {list(df_raw_trades.columns)}")
                # Skip creating trade keys for raw trades if we can't identify the side column
                df_raw_trades['trade_key'] = ''
            
            if 'trade_key' not in df_raw_trades.columns:
                df_raw_trades['trade_key'] = (
                    df_raw_trades['Symbol'].astype(str) + '_' + 
                    df_raw_trades['Date'].astype(str) + '_' + 
                    df_raw_trades['Time'].astype(str) + '_' + 
                    df_raw_trades[side_col].astype(str) + '_' + 
                    df_raw_trades[qty_col].astype(str) + '_' + 
                    df_raw_trades['Price'].astype(str)
                )
        
        # Create unique identifier for existing consolidated trades
        if not df_consolidated.empty:
            # Check for Side column more carefully
            if 'Side' in df_consolidated.columns:
                side_col = 'Side'
            elif 'Type' in df_consolidated.columns:
                side_col = 'Type'
            else:
                print(f"⚠️ Warning: No 'Side' or 'Type' column found in Consolidated Trades. Available columns: {list(df_consolidated.columns)}")
                df_consolidated['trade_key'] = ''
            
            if 'trade_key' not in df_consolidated.columns and side_col:
                # Use the correct column name - should match the consolidated_trade_key format
                # consolidated_trade_key = f"{trade['Symbol']}_{trade['Date']}_{trade['Side']}"
                date_col = 'Processed' if 'Processed' in df_consolidated.columns else 'Date'
                df_consolidated['trade_key'] = (
                    df_consolidated['Symbol'].astype(str) + '_' + 
                    df_consolidated[date_col].astype(str) + '_' + 
                    df_consolidated[side_col].astype(str)
                )
        
        # Track new trades for all sheets
        new_raw_trades = []
        
        # Group consolidated trades by symbol, date, and side for the consolidated sheet
        consolidated_by_day = {}
        
        for trade in consolidated_trades:
            # Create trade keys
            position_trade_key = (
                f"{trade['Symbol']}_{trade['Quantity']}_{trade['Side']}_"
                f"{trade['Price']}_{trade['Time']}_{trade['Date']}"
            )
            
            raw_trade_key = (
                f"{trade['Symbol']}_{trade['Date']}_{trade['Time']}_{trade['Side']}_"
                f"{trade['Quantity']}_{trade['Price']}"
            )
            
            consolidated_trade_key = f"{trade['Symbol']}_{trade['Date']}_{trade['Side']}"
            
            # Add to raw trades sheet if not already exists
            if df_raw_trades.empty or raw_trade_key not in df_raw_trades['trade_key'].values:
                new_raw_trade = {
                    "Symbol": trade['Symbol'],
                    "Quantity": trade['Quantity'],
                    "Side": trade['Side'],  # Add the missing Side column
                    "Price": trade['Price'],
                    "Time": trade['Time'],
                    "Date": pd.to_datetime(trade['Date']).strftime('%Y-%m-%d')
                }
                new_raw_trades.append(new_raw_trade)
            
            # Group for consolidated trades sheet (by symbol, date, side)
            if consolidated_trade_key not in consolidated_by_day:
                consolidated_by_day[consolidated_trade_key] = {
                    "Symbol": trade['Symbol'],
                    "Date": pd.to_datetime(trade['Date']).strftime('%Y-%m-%d'),
                    "Side": trade['Side'],
                    "total_qty": 0,
                    "total_value": 0,
                    "time": trade['Time']  # Make sure this is the actual time, not None
                }
            
            group = consolidated_by_day[consolidated_trade_key]
            group['total_qty'] += trade['Quantity']
            group['total_value'] += trade['Quantity'] * trade['Price']
            
            # Update time based on trade side - ensure we're not getting None/NaN
            current_time = trade['Time']
            if current_time and str(current_time) != 'nan':
                # LONG: keep earliest time
                # SHORT: keep latest time
                if trade['Side'] == 'LONG':
                    if group['time'] and str(group['time']) != 'nan':
                        group['time'] = min(group['time'], current_time)
                    else:
                        group['time'] = current_time
                else:  # SHORT
                    if group['time'] and str(group['time']) != 'nan':
                        group['time'] = max(group['time'], current_time)
                    else:
                        group['time'] = current_time
            
            # Note: Trades sheet (position tracking) is built by match_trades_fifo
            # below — it handles entries, FIFO closes, and split-on-partial. We
            # intentionally do not pre-append entries here.
        
        # Create new consolidated trades for the consolidated sheet
        new_consolidated_trades = []
        for key, group in consolidated_by_day.items():
            # Check if this consolidated trade already exists
            if df_consolidated.empty or key not in df_consolidated['trade_key'].values:
                # Fix the average price calculation for SHORT positions
                if group['total_qty'] != 0:
                    avg_price = abs(group['total_value'] / group['total_qty'])
                else:
                    avg_price = 0
                    
                new_consolidated_trade = {
                    "Symbol": group['Symbol'],
                    "Quantity": group['total_qty'],
                    "Side": group['Side'],
                    "Avg_Price": avg_price,
                    "Time": group['time'],  # Make sure time is included
                    "Processed": pd.to_datetime(group['Date']).strftime('%Y-%m-%d')
                }
                
                new_consolidated_trades.append(new_consolidated_trade)
        
        if new_raw_trades:
            df_new_raw = pd.DataFrame(new_raw_trades)
            df_raw_trades = pd.concat([df_raw_trades, df_new_raw], ignore_index=True)

        if new_consolidated_trades:
            df_new_consolidated = pd.DataFrame(new_consolidated_trades)
            df_consolidated = pd.concat([df_consolidated, df_new_consolidated], ignore_index=True)

        # Drop temporary trade_key columns before saving
        df_master = df_master.drop('trade_key', axis=1, errors='ignore')
        df_raw_trades = df_raw_trades.drop('trade_key', axis=1, errors='ignore')
        df_consolidated = df_consolidated.drop('trade_key', axis=1, errors='ignore')

        # Run FIFO: matches new closes against existing open lots (handling
        # legacy partials and split-on-close) and creates new open rows for
        # the remainder. Operates on the existing master rows directly.
        df_master = match_trades_fifo(df_master, consolidated_trades)
        
        # Sort raw trades sheet by date and time
        if not df_raw_trades.empty:
            # Convert both date and time to string before concatenating, using mixed format for flexibility
            df_raw_trades['datetime'] = pd.to_datetime(df_raw_trades['Date'].astype(str) + ' ' + df_raw_trades['Time'].astype(str), format='mixed', errors='coerce')
            df_raw_trades = df_raw_trades.sort_values('datetime').drop('datetime', axis=1)
        
        # Sort consolidated trades sheet by date
        # Create unique identifier for existing consolidated trades
        if not df_consolidated.empty:
            # Check for Side column more carefully
            if 'Side' in df_consolidated.columns:
                side_col = 'Side'
            elif 'Type' in df_consolidated.columns:
                side_col = 'Type'
            else:
                print(f"⚠️ Warning: No 'Side' or 'Type' column found in Consolidated Trades. Available columns: {list(df_consolidated.columns)}")
                df_consolidated['trade_key'] = ''
            
            if 'trade_key' not in df_consolidated.columns and side_col:
                # Use the correct column name - should match the consolidated_trade_key format
                # consolidated_trade_key = f"{trade['Symbol']}_{trade['Date']}_{trade['Side']}"
                date_col = 'Processed' if 'Processed' in df_consolidated.columns else 'Date'
                df_consolidated['trade_key'] = (
                    df_consolidated['Symbol'].astype(str) + '_' + 
                    df_consolidated[date_col].astype(str) + '_' + 
                    df_consolidated[side_col].astype(str)
                )
        # Save updated master file with proper sheet names
        with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
            df_master.to_excel(writer, sheet_name='Trades', index=False)  # Position tracking
            df_raw_trades.to_excel(writer, sheet_name='Raw Trades', index=False)  # All individual trades
            df_consolidated.to_excel(writer, sheet_name='Consolidated Trades', index=False)  # Daily consolidation
        
        parent_dir = os.path.basename(os.path.dirname(master_file))
        filename = os.path.basename(master_file)
        print(f"✅ Updated {parent_dir}/{filename}:")
        print(f"   - 'Trades' sheet now {len(df_master)} rows (FIFO matched)")
        print(f"   - Added {len(new_raw_trades)} new trades to 'Raw Trades' sheet")
        print(f"   - Added {len(new_consolidated_trades)} new entries to 'Consolidated Trades' sheet")
        print(f"📊 Final data after processing:")
        print(f"   - Trades sheet: {len(df_master)} rows")
        print(f"   - Raw Trades sheet: {len(df_raw_trades)} rows")
        print(f"   - Consolidated Trades sheet: {len(df_consolidated)} rows")
        
    except Exception as e:
        print(f"❌ Error updating master sheet: {str(e)}")
        import traceback
        traceback.print_exc()

def manage_processed_files(folder_path, pdf_file=None, check_only=False):
    """Track processed PDF files using a JSON file"""
    tracking_file = os.path.join(folder_path, PROCESSED_FILE)  # Use test file if in test mode
    
    # Load existing processed files
    if os.path.exists(tracking_file):
        with open(tracking_file, 'r') as f:
            processed_files = json.load(f)
    else:
        processed_files = []
    
    if check_only:
        return processed_files
    
    # Add new file and save
    if pdf_file and pdf_file not in processed_files:
        processed_files.append(pdf_file)
        with open(tracking_file, 'w') as f:
            json.dump(processed_files, f, indent=2)
    
    return processed_files

def reset_test_files(folder_path):
    """Reset test files before running script"""
    if TEST_MODE:
        BASE_PATH = BASE_PATH_TRADES
        
        # Reset processed files JSON
        test_json_path = os.path.join(folder_path, PROCESSED_FILE)
        with open(test_json_path, 'w') as f:
            json.dump([], f)
        print("🔄 Reset processed files tracking")
        
        # Create empty test master copy if it doesn't exist
        test_master_path = os.path.join(BASE_PATH, MASTER_FILE)
        if not os.path.exists(test_master_path):
            df_empty = pd.DataFrame(columns=[
                "Symbol", "Qty", "Side", "Entry Price", "Entry Time", 
                "Entry Date", "Notes", "Exit Qty", "Exit Price", 
                "Exit Time", "Exit Date"
            ])
            df_empty.to_excel(test_master_path, index=False)
        print("🔄 Reset master copy test file")

def reset_master_sheet():
    """Reset all spreadsheets and processed files tracking"""
    try:
        BASE_PATH = BASE_PATH_TRADES
        master_file = os.path.join(BASE_PATH, MASTER_FILE)
        backup_file = os.path.join(BASE_PATH, MASTER_BACKUP)
        
        print(f"\n🔄 Resetting all spreadsheets and processed files...")
        
        # Create backup before reset
        if os.path.exists(master_file):
            print(f"📑 Creating backup before reset...")
            with pd.ExcelFile(master_file) as xls:
                all_sheets = {}
                for sheet_name in xls.sheet_names:
                    all_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
            
            # Save backup with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            reset_backup = os.path.join(BASE_PATH, f"master_pre_reset_{timestamp}.xlsx")
            with pd.ExcelWriter(reset_backup, engine='openpyxl') as writer:
                for sheet_name, df in all_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Pre-reset backup created: {os.path.basename(reset_backup)}")
        
        # Create fresh empty sheets with headers only
        df_master = pd.DataFrame(columns=[
            "Symbol", "Qty", "Side", "Entry Price", "Entry Time", 
            "Entry Date", "Notes", "Exit Qty", "Exit Price", 
            "Exit Time", "Exit Date"
        ])
        
        df_raw_trades = pd.DataFrame(columns=[
            "Symbol", "Quantity", "Side", "Price", "Time", "Date"
        ])
        
        df_consolidated = pd.DataFrame(columns=[
            "Symbol", "Quantity", "Side", "Avg_Price", "Time", "Processed"
        ])
        
        # Save empty master file
        with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
            df_master.to_excel(writer, sheet_name='Trades', index=False)
            df_raw_trades.to_excel(writer, sheet_name='Raw Trades', index=False)
            df_consolidated.to_excel(writer, sheet_name='Consolidated Trades', index=False)
        
        # Reset all processed_files.json in subdirectories
        reset_count = 0
        for root, dirs, files in os.walk(BASE_PATH):
            processed_file = os.path.join(root, "processed_files.json")
            if os.path.exists(processed_file):
                # Clear the processed files list
                with open(processed_file, 'w') as f:
                    json.dump([], f, indent=2)
                reset_count += 1
                rel_path = os.path.relpath(root, BASE_PATH)
                print(f"   📂 Reset processed files in: {rel_path}")
        
        # Remove backup file if it exists
        if os.path.exists(backup_file):
            os.remove(backup_file)
        
        print(f"✅ Reset complete!")
        print(f"   - Cleared all 3 spreadsheet tabs (keeping headers)")
        print(f"   - Reset {reset_count} processed_files.json files")
        print(f"   - All trade data has been cleared")
        print(f"   - You can now reprocess folders from scratch")
        
    except Exception as e:
        print(f"❌ Error during reset: {str(e)}")

JOURNAL_FILE = "Trades.xlsx"
JOURNAL_SHEET = "Trades"
# Columns FIFO automation manages directly. All other columns (Setup,
# Entry Notes, Stop Price, Target Price, P/L formulas, etc.) are preserved
# untouched on existing rows; on new rows they are left blank, but cells
# containing formulas are copied from the previous row so drag-down stays
# intact.
_JOURNAL_FIFO_COLS = {
    "Symbol", "Qty", "Side", "Entry Price", "Entry Time", "Entry Date",
    "Exit Qty", "Exit Price", "Exit Time", "Exit Date",
}


def _journal_normalize_date(v):
    """Normalize an Excel cell value to an ISO date string for FIFO comparison."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if hasattr(v, 'strftime'):  # date or Timestamp
        try:
            return v.strftime('%Y-%m-%d')
        except Exception:
            pass
    return str(v).split(' ')[0]


def _journal_normalize_time(v):
    if v is None or v == '':
        return None
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%H:%M:%S')
        except Exception:
            pass
    return str(v)


def update_trades_journal(consolidated_trades, folder_path):
    """Apply FIFO matching to user's Trades.xlsx journal in place, preserving
    formulas and custom columns (Setup, Notes, Stop/Target, P/L, etc.).

    For existing rows: closes are written into Exit Qty/Price/Time/Date.
    Open rows that receive a partial close are split into a closed lot row
    (appended) plus the original row with Qty reduced by the fill.
    Brand-new entries with no opposite-side open lot become appended rows.
    """
    import shutil
    from openpyxl import load_workbook
    from openpyxl.formula.translate import Translator
    from openpyxl.worksheet.formula import ArrayFormula

    journal_path = os.path.join(BASE_PATH_TRADES, JOURNAL_FILE)
    if not os.path.exists(journal_path):
        print(f"⚠️  {JOURNAL_FILE} not found at {journal_path}; skipping journal update")
        return

    print(f"\n📓 Updating {JOURNAL_FILE} journal...")

    # Backup
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(BASE_PATH_TRADES, f'Trades_backup_{ts}.xlsx')
    shutil.copy(journal_path, backup_path)
    print(f"📑 Backup: {os.path.basename(backup_path)}")

    wb = load_workbook(journal_path)
    if JOURNAL_SHEET not in wb.sheetnames:
        print(f"⚠️ '{JOURNAL_SHEET}' sheet not found in {JOURNAL_FILE}; skipping")
        return
    ws = wb[JOURNAL_SHEET]

    # Map first occurrence of each header to its column index. Duplicates
    # (file has a stray second 'Symbol' column with broken array formulas)
    # are left strictly alone.
    headers = {}
    duplicates = set()
    for idx, c in enumerate(ws[1], 1):
        h = c.value
        if h is None or h == '':
            continue
        if h in headers:
            duplicates.add(idx)
            continue
        headers[h] = idx
    required = ['Symbol', 'Qty', 'Side', 'Entry Price', 'Entry Time', 'Entry Date',
                'Exit Qty', 'Exit Price', 'Exit Time', 'Exit Date']
    missing = [c for c in required if c not in headers]
    if missing:
        print(f"⚠️ Missing required columns in {JOURNAL_SHEET} sheet: {missing}; skipping")
        return

    # Find true last data row (last row with a non-empty Symbol).
    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=headers['Symbol']).value not in (None, ''):
            last_row = r

    # Read existing rows into FIFO records.
    fifo_rows = []
    for r in range(2, last_row + 1):
        sym = ws.cell(row=r, column=headers['Symbol']).value
        if sym in (None, ''):
            continue
        rec = {
            'Symbol': sym,
            'Qty': ws.cell(row=r, column=headers['Qty']).value,
            'Side': ws.cell(row=r, column=headers['Side']).value,
            'Entry Price': ws.cell(row=r, column=headers['Entry Price']).value,
            'Entry Time': _journal_normalize_time(ws.cell(row=r, column=headers['Entry Time']).value),
            'Entry Date': _journal_normalize_date(ws.cell(row=r, column=headers['Entry Date']).value),
            'Notes': '',
            'Exit Qty': ws.cell(row=r, column=headers['Exit Qty']).value,
            'Exit Price': ws.cell(row=r, column=headers['Exit Price']).value,
            'Exit Time': _journal_normalize_time(ws.cell(row=r, column=headers['Exit Time']).value),
            'Exit Date': _journal_normalize_date(ws.cell(row=r, column=headers['Exit Date']).value),
            '_xlsx_row': r,
        }
        fifo_rows.append(rec)

    pre_count = len(fifo_rows)
    # Snapshot the FIFO-managed fields of each existing row so we can detect
    # which ones FIFO actually mutated (avoid pointless writes that could
    # nudge cell formatting).
    snapshot = {
        id(r): (r.get('Qty'), r.get('Exit Qty'), r.get('Exit Price'),
                r.get('Exit Time'), r.get('Exit Date'))
        for r in fifo_rows
    }
    snap_by_xrow = {r['_xlsx_row']: snapshot[id(r)] for r in fifo_rows}

    out_rows = _match_trades_fifo_records(fifo_rows, consolidated_trades)

    def is_dirty(rec):
        xrow = rec.get('_xlsx_row')
        if xrow not in snap_by_xrow:
            return True
        before = snap_by_xrow[xrow]
        after = (rec.get('Qty'), rec.get('Exit Qty'), rec.get('Exit Price'),
                 rec.get('Exit Time'), rec.get('Exit Date'))
        return before != after

    # Categorize output rows for write-back:
    #   - in_place_updates: existing rows whose FIFO fields changed
    #   - split_children:   newly-created lots (must be appended, inherit
    #                       static cols + formulas from parent's xlsx row)
    #   - new_appends:      brand-new entries (no parent row; only formulas
    #                       are inherited from the prior journal row)
    in_place_updates = []
    split_children = []
    new_appends = []
    for r in out_rows:
        if r.get('_split_child'):
            split_children.append(r)
        elif r.get('_xlsx_row'):
            if is_dirty(r):
                in_place_updates.append(r)
        else:
            new_appends.append(r)

    # Helpers
    def journal_side(side_internal, original=None):
        """Convert internal BUY/SELL to LONG/SHORT, or fall back to whatever
        the user had on the original row."""
        if original is not None and original.get('Side') in ('LONG', 'SHORT'):
            # Prefer to keep whatever convention the user already uses.
            pass
        if side_internal == 'BUY':
            return 'LONG'
        if side_internal == 'SELL':
            return 'SHORT'
        return side_internal

    def write_fifo_fields(target_row, rec):
        """Write only the FIFO-managed fields onto the given xlsx row."""
        ws.cell(row=target_row, column=headers['Symbol']).value = rec['Symbol']
        # Store Qty as positive integer in journal (user convention: LONG/SHORT
        # carries the direction; Qty is magnitude).
        qv = rec.get('Qty')
        if qv is not None:
            ws.cell(row=target_row, column=headers['Qty']).value = abs(int(qv))
        ws.cell(row=target_row, column=headers['Side']).value = journal_side(rec.get('Side'))
        ws.cell(row=target_row, column=headers['Entry Price']).value = rec.get('Entry Price')
        ws.cell(row=target_row, column=headers['Entry Time']).value = rec.get('Entry Time')
        ed = rec.get('Entry Date')
        ws.cell(row=target_row, column=headers['Entry Date']).value = (
            datetime.strptime(ed, '%Y-%m-%d') if isinstance(ed, str) else ed
        )

        eq = rec.get('Exit Qty')
        if eq is None or (isinstance(eq, float) and pd.isna(eq)):
            ws.cell(row=target_row, column=headers['Exit Qty']).value = None
        else:
            ws.cell(row=target_row, column=headers['Exit Qty']).value = abs(int(eq)) if isinstance(eq, (int, float)) and not pd.isna(eq) else eq
        ws.cell(row=target_row, column=headers['Exit Price']).value = rec.get('Exit Price')
        ws.cell(row=target_row, column=headers['Exit Time']).value = rec.get('Exit Time')
        xd = rec.get('Exit Date')
        if xd is None or (isinstance(xd, float) and pd.isna(xd)):
            ws.cell(row=target_row, column=headers['Exit Date']).value = None
        else:
            ws.cell(row=target_row, column=headers['Exit Date']).value = (
                datetime.strptime(xd, '%Y-%m-%d') if isinstance(xd, str) else xd
            )

    # For each column, find the most recent row (<= last_row) that contains
    # a formula in that column. Used as the translation source so we always
    # pull from a closed-lot template even when nearby rows are open.
    formula_src_for_col = {}
    for col_idx in range(1, ws.max_column + 1):
        if col_idx in duplicates:
            continue
        for r in range(last_row, 1, -1):
            v = ws.cell(row=r, column=col_idx).value
            if isinstance(v, str) and v.startswith('='):
                formula_src_for_col[col_idx] = r
                break

    def _translate_formula_into(col_idx, dst_row):
        """Pull the latest formula in col_idx, translate to dst_row, write."""
        src_r = formula_src_for_col.get(col_idx)
        if not src_r or src_r == dst_row:
            return
        src_cell = ws.cell(row=src_r, column=col_idx)
        v = src_cell.value
        if not (isinstance(v, str) and v.startswith('=')):
            return
        try:
            new_f = Translator(v, origin=src_cell.coordinate).translate_formula(
                ws.cell(row=dst_row, column=col_idx).coordinate
            )
            ws.cell(row=dst_row, column=col_idx).value = new_f
        except Exception:
            pass

    def copy_static_columns(src_row, dst_row):
        """For split-children: copy non-FIFO columns from parent row.
        - Plain text/number values are copied as-is (Setup, Notes, etc.).
        - Formula cells are translated to dst_row.
        - Cells blank on parent get a formula filled in if any prior row has
          one for that column (so closed split-children get P/L formulas
          even when parent was an open row with blank P/L).
        - Array formulas and duplicate-header columns are skipped.
        """
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in duplicates:
                continue
            header_name = ws.cell(row=1, column=col_idx).value
            if header_name in _JOURNAL_FIFO_COLS:
                continue
            src_cell = ws.cell(row=src_row, column=col_idx)
            v = src_cell.value
            if isinstance(v, ArrayFormula):
                continue
            if isinstance(v, str) and v.startswith('='):
                try:
                    new_f = Translator(v, origin=src_cell.coordinate).translate_formula(
                        ws.cell(row=dst_row, column=col_idx).coordinate
                    )
                    ws.cell(row=dst_row, column=col_idx).value = new_f
                except Exception:
                    pass
            elif v is not None:
                ws.cell(row=dst_row, column=col_idx).value = v
            else:
                _translate_formula_into(col_idx, dst_row)

    def add_formulas_for_close(dst_row):
        """For newly-closed rows (in-place close transition or split-child):
        fill in formula columns that are currently blank, pulled from the
        latest-formula source row for each column."""
        for col_idx in formula_src_for_col:
            header_name = ws.cell(row=1, column=col_idx).value
            if header_name in _JOURNAL_FIFO_COLS:
                continue
            target = ws.cell(row=dst_row, column=col_idx)
            if target.value in (None, ''):
                _translate_formula_into(col_idx, dst_row)

    next_append = last_row + 1
    n_updates = n_splits = n_new = 0

    for rec in in_place_updates:
        xrow = rec['_xlsx_row']
        before = snap_by_xrow.get(xrow, (None,) * 5)
        was_open_pre = _is_blank(before[3]) and _is_blank(before[4])  # Exit Time + Exit Date blank
        write_fifo_fields(xrow, rec)
        # If this row just transitioned to closed (Exit Date now filled),
        # add P/L (and any other formula columns) that were blank.
        if was_open_pre and not _is_blank(rec.get('Exit Date')):
            add_formulas_for_close(xrow)
        n_updates += 1

    for child in split_children:
        parent_row = child['_xlsx_row']
        copy_static_columns(parent_row, next_append)
        write_fifo_fields(next_append, child)
        # Split-children are always closed lots; ensure formulas exist.
        add_formulas_for_close(next_append)
        next_append += 1
        n_splits += 1

    for rec in new_appends:
        write_fifo_fields(next_append, rec)
        # Open-and-closed-in-same-run case (e.g. day-trade): Exit Date is
        # already filled, so populate formula columns. Otherwise leave
        # blank — user fills them on close (matching the drag-down workflow).
        if not _is_blank(rec.get('Exit Date')):
            add_formulas_for_close(next_append)
        next_append += 1
        n_new += 1

    wb.save(journal_path)
    print(
        f"✅ Updated {JOURNAL_FILE}: "
        f"{n_updates} row(s) updated in place, "
        f"{n_splits} split-child row(s) appended, "
        f"{n_new} new entry row(s) appended."
    )
    print(f"   - Journal had {pre_count} data rows; now has {next_append - 2} data rows.")


def process_folder(date_str):
    """Process a single folder based on date string"""
    try:
        folder_path = get_folder_path(date_str)
        print(f"\n📁 Processing folder: {os.path.basename(folder_path)}")
        
        # Reset test files if in test mode
        if TEST_MODE:
            reset_test_files(folder_path)
        
        # Get all trades from PDFs in the folder
        all_trades = gather_all_trades(folder_path)
        
        if not all_trades:
            print("No new trades found to process.")
            return
        
        # Consolidate trades by symbol and date
        consolidated_trades = consolidate_trades(all_trades)
        
        print(f"\n📊 Trade Summary:")
        print(f"   - Total individual trades: {len(all_trades)}")
        print(f"   - Consolidated trades: {len(consolidated_trades)}")
        
        # Update master sheet with consolidated trades (audit trail).
        update_master_sheet(consolidated_trades, folder_path)

        # Update the user's Trades.xlsx journal with FIFO matching applied
        # in-place (preserves formulas and custom columns).
        update_trades_journal(consolidated_trades, folder_path)

        # Check and display open positions
        check_open_positions(folder_path)
        
    except (FileNotFoundError, ValueError) as e:
        print(f"❌ Error: {str(e)}")
    except Exception as e:
        print(f"❌ Unexpected error processing folder: {str(e)}")

def main():
    print("=" * 60)
    print("📊 TRADE LOG FORMATTER")
    print("=" * 60)
    
    # Get current month/year as default
    current_month_year = datetime.now().strftime("%m.%Y")
    
    choice = input(f"\n'RESET' or enter a date (default: {current_month_year}): ").strip()
    
    if choice == 'RESET':
        confirm = input("⚠️  This will DELETE ALL trade data. Type 'y' to confirm: ").strip()
        if confirm == 'y':
            reset_master_sheet()
        else:
            print("❌ Reset cancelled")
    else:
        # Use current month/year if no input provided
        date_to_process = choice if choice else current_month_year
        print(f"📅 Processing folder: {date_to_process}")
        process_folder(date_to_process)

if __name__ == "__main__":
    main()